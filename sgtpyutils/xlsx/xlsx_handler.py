import inspect
from .openpyxl_extend import *
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
# 用于对Excel文件进行操作

import os


class XlsxHandler:
    use_new_sheet = True
    wb: openpyxl.Workbook = None

    def __init__(self, target_file: str):
        self.target_file = target_file
        if os.path.exists(self.target_file):
            self.wb = openpyxl.load_workbook(
                filename=self.target_file)  # 缓存excel文件对象
        else:
            print(f'warn:file not exist:{self.target_file},use in memory')
            self.wb = openpyxl.Workbook()


    def read(self, sheet: str = None, headers: dict = None, mapper: dict = None, value_convert: dict = None):
        '''
        将表格转换成为list类型方便后续操作
        sheet		:string	需要读取的表名称,默认读取第一个表
        headers	:dict	需要读取的表头名称,默认读全部列
        mapper    :dict   将原始表头转换为对应的key
        value_convert :dict 实现excel输入项的值转换
        '''
        sh = self.wb[0] if sheet is None else self.wb[sheet]
        sh_headers = self.__get_headers_from_sheet(sh, headers, mapper)
        return self.__read_data(sh, sh_headers, value_convert)

    def __read_data(self, sh: Worksheet, headers: dict, value_convert: dict):
        if not value_convert:
            value_convert = {}
        any_data = {}
        result = [{} for x in range(sh.max_row)]  # 初始化数据
        for row in range(1, sh.max_row+1):
            for col in headers:
                item = headers[col]
                v = sh.cell(row+1, item['index']).value  # 从第2行开始读取
                name = item['name']
                if v:
                    any_data[row] = True
                    # 将当前行数据入内存
                    result[row-1][name] = XlsxHandler.__convert_data(
                        v, value_convert.get(name))
        # 排除无效行
        return [x for index, x in enumerate(result) if index+1 in any_data]

    def __convert_data(data: str, converter):
        if not converter or data is None:
            return data
        if inspect.isfunction(converter):
            return converter(data)
        if converter == 'list':
            return data.split(',')
        return data

    

    def write(self, data: dict, sheet: str = None, append: bool = False):
        '''
        将数据写入到表格中
        sheet		:string	需要读取的表名称,创建下一个表
        data      :Object 需要存入的数据
        '''
        sh, is_new_sheet = get_sheet_with_create(
            self.wb, sheet, put_index=0, use_new_sheet=not append and self.use_new_sheet)
        headers = self.__get_headers_from_data(data)
        self.__build_headers(sh, headers)
        self.__build_values(sh, data, headers, append)
        self.wb.save(self.target_file)

    def __build_values(self, sh: Worksheet, data: dict, headers: list, append: bool = False):
        current_index = 1 if not append else sh.max_row
        current_index += 1  # 从下一行开始
        for index, item in enumerate(data):
            for h_index, header in enumerate(headers):
                if not header in item:
                    continue
                v = XlsxHandler.__to_value(item[header])
                sh.cell(index+current_index, h_index + 1).value = v

    def __to_value(value) -> str:
        if isinstance(value, int) or isinstance(value, float) or isinstance(value, str):
            return value
        if isinstance(value, list):
            return ','.join([str(x) for x in value])
        return str(value)

    def __build_headers(self, sh, headers):
        for index, h in enumerate(headers):
            sh.cell(1, index+1).value = h

    def __get_headers_from_data(self, data):
        result = {}
        for item in data:
            for k in item:
                result[k] = True
        return list(result)

    def __get_headers_from_sheet(self, sh, headers=None, mapper=None):
        if mapper is None:
            mapper = {}
        sh_headers = {}  # 加载header的列索引
        for col in range(1, sh.max_column+1):
            h = sh.cell(1, col).value
            if h is None:
                continue
            if not headers is None and not h in headers:
                continue
            if h in sh_headers:
                print(f'warn:sheet_{sh.title}_{h} is duplicated')
            if h in mapper:
                h = mapper[h]
            sh_headers[h] = {
                'index': col,
                'name': h
            }
        return sh_headers
